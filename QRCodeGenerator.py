import ssl
import json
import qrcode
import smtplib
import openpyxl
from lib import *
from os import system
from time import time, strftime, gmtime
from email.message import EmailMessage

begin_time = time()
system('cls')
print('==> {:.2f} %'.format(0))

json_data = {}
with open("settings.json", "r") as data_file:
    settings = json.load(data_file)
with open(settings["app_password_file"], "r") as apppass:
    email_password = apppass.readline().strip()

# To open the workbook workbook object is created
wb_obj = openpyxl.load_workbook(settings["excel_file"])
# Get workbook active sheet object from the active attribute
sheet = wb_obj.active

# Set the subject and body of the email (event name for exp)
subject = 'The Purge has began!'

for row_id in range(2, sheet.max_row+1):
    id_cell = sheet.cell(row=row_id, column=1).value
    prtc_id = tstamp_to_id(str(id_cell))
    prtc_first_name = sheet.cell(row=row_id, column=2).value
    prtc_last_name = sheet.cell(row=row_id, column=3).value
    prtc_email = sheet.cell(row=row_id, column=4).value
    # prtc_number = sheet.cell(row=row_id, column=5).value
    json_data[prtc_id] = [prtc_first_name,
                          prtc_last_name, prtc_email, prtc_id, []]

    # Data to be encoded
    # data = 'QR Code using make() function'
    data = f"{prtc_id}-{prtc_first_name}-{prtc_last_name}"

    # Encoding data using make() function
    img = qrcode.make(data)
    # TODO: check if the folder exist
    # Saving as an image file
    img.save(f'Generated QR code/id-{row_id-1}.png')

    body = f"""
    Good evening {prtc_first_name} {prtc_last_name}
    we hope that all is good
    
    that's a test of the auto emails script
    your QR code is attached?
    
    
    Best regards,
    GDG
    """
    email_receiver = prtc_email
    em = EmailMessage()
    em['From'] = settings["email_sender"]
    em['To'] = email_receiver
    em['Subject'] = subject
    em.set_content(body)
    with open(f"Generated QR code\id-{row_id-1}.png", 'rb') as fp:
        img_data = fp.read()
    em.add_attachment(img_data, maintype='image', subtype='png')

    # Sending the email: Add SSL (layer of security)
    context = ssl.create_default_context()
    # Log in and send the email
    with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
        smtp.login(settings["email_sender"], email_password)
        smtp.sendmail(settings["email_sender"], email_receiver, em.as_string())

    system('cls')
    print('==> {:.2f} %'.format((row_id)*100/sheet.max_row))

# TODO: check if the file exist
with open(settings["participantes_data_file"], "w") as prtc_data_file:
    prtc_data_file.write(json.dumps(json_data, indent=4))

print('>> All is Done Here xD')
print('>> Task finished in: {}'.format(
    strftime("%H:%M:%S", gmtime(time()-time()+1000))))
