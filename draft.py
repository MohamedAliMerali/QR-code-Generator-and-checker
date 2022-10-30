# 10/2/2022 21:01:53
# https://myaccount.google.com/apppasswords?rapt=AEjHL4MfNrq7onFGWSIELHL9B7cQHFwBJOmP1Tk4aLi1nBZM7BUYmQ3EEZOPCw93queTgJTp-Nkg1hD7_64dNZcsHS_TWQLYPQ

# participantes_file_path = "participante.xlsx"
# participantes_file = open(participantes_file_path)
# participantes_csv_file = csv.reader(participantes_file)

# rows = []
# for row in participantes_csv_file:
#     rows.append(row)

# #################################################################

# # Define variable to load the dataframe
# dataframe = openpyxl.load_workbook("participante.xlsx")

# # Define variable to read sheet
# dataframe1 = dataframe.active

# # Iterate the loop to read the cell values
# for row in range(0, dataframe1.max_row):
#     for col in dataframe1.iter_cols(1, dataframe1.max_column):
#         print(col[row].value)
# #################################################################

# # Python3 code to select
# # data from excel
# import xlwings as xw
# import numpy as np
# import qrcode
# import smtplib
# import ssl
# from email.message import EmailMessage
# from os import system

# # TODO: make generate ID function

# with open("apppassword.txt", "r") as apppass:
#     lines = apppass.readline().strip()

# # Define email sender and receiver
# email_sender = 'merali.med@gmail.com'
# email_password = lines
# # Set the subject and body of the email (event name for exp)
# subject = 'The Purge has began!'

# # Specifying a sheet
# ws = xw.Book("participante.xlsx").sheets['Sheet1']

# header = np.array(ws.range("A1:F1").value)
# ids = np.array(ws.range("A2:A6").value)
# users_informations = np.array(ws.range("B2:F6").value)


# for index, informations in enumerate(users_informations):
#     system('cls')
#     print('==> {:.2f} %'.format((index+1)*100/len(users_informations)))
#     # Data to be encoded
#     # data = 'QR Code using make() function'
#     data = f"{informations[0]}-{informations[1]}-{informations[2]}"

#     # Encoding data using make() function
#     img = qrcode.make(data)
#     # TODO: check if the folder exist
#     # Saving as an image file
#     img.save(f'Generated QR code/id-{index}.png')

#     body = f"""
#     Good evening {informations[0]} {informations[1]}
#     we hope that all is good
    
#     that's a test of the auto emails script
#     in the next email you will receive you QR code
    
    
#     Best regards,
#     GDG
#     """
#     email_receiver = informations[2]
#     em = EmailMessage()
#     em['From'] = email_sender
#     em['To'] = email_receiver
#     em['Subject'] = subject
#     em.set_content(body)
#     with open("Generated QR code\id-0.png", 'rb') as fp:
#         img_data = fp.read()
#     em.add_attachment(img_data, maintype='image',
#                       subtype='png')

#     # Sending the email:
#     # Add SSL (layer of security)
#     context = ssl.create_default_context()
#     # Log in and send the email
#     with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
#         smtp.login(email_sender, email_password)
#         smtp.sendmail(email_sender, email_receiver, em.as_string())


# print('All is Done Here xD')


import openpyxl


def tstamp_to_id(tstamp):
    date, time = tstamp.split(" ")
    date = date.split("/")
    time = time.split(":")
    date.pop()
    date.extend(time)
    return "".join(date)


excel_path = "participante.xlsx"

# To open the workbook workbook object is created
wb_obj = openpyxl.load_workbook(excel_path)
# Get workbook active sheet object from the active attribute
sheet = wb_obj.active

print(sheet.cell(row=2, column=1).value)
